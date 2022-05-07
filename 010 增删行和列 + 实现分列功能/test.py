import openpyxl
import tools


def text_to_columns(index: int, delimiter :str , number:int ,worksheet):
    '''

    :param index:
    :param delimeter:
    :param numner:
    :param worksheet:
    :return:


    '''
    max_row = worksheet.max_row
    # 需要准备空列
    worksheet.insert_cols(index + 1, number - 1)
    # 填充表头

    for i in range(index, index + number):
        ws.cell(row=1, column=i, value=str(i - index + 1) + '级科目')
    # 针对范围内的值进行分割,并进行填充
    # 分割每一个单元格
    for i in range(2, max_row):

        split_cell = ws.cell(row=i, column=index).value.split(delimiter)

        # 判断split_cell的长度和和number的关系
        if len(split_cell) > number:

            for j in range(0, number):
                ws.cell(row=i, column=index + j, value=split_cell[j].strip())

        else:
            j = 0
            for each_content in split_cell:
                ws.cell(row=i, column=index + j, value=each_content.strip())
                j = j + 1

    return worksheet


def process_worksheet(worksheet):
    tools.fill_column(5,worksheet)
    tools.fill_column(6, worksheet)
    worksheet.delete_cols(15,18)
    worksheet.delete_cols(10,3)
    worksheet.delete_cols(8)
    worksheet.delete_cols(1, 4)
    text_to_columns(4, '-', 3, worksheet)
    return worksheet

if __name__ == '__main__':
    wb = openpyxl.load_workbook('kisdocument.xlsx')
    ws = wb.active
    process_worksheet(ws)
    wb.save('result.xlsx')
