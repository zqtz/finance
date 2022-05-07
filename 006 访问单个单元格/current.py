import openpyxl
import tools

# ws = tools.open_xlsx_file('cell.xlsx')
# cell:openpyxl.cell.cell.Cell = ws['B2']
# print(type(cell))
# print(cell.col_idx)
# print(cell.column)
# print(cell.row)
# print(cell.coordinate)
# print(cell.column_letter)
# print(cell.encoding)
# print(cell.is_date)
# print(cell.value)

def rotate(row:int,column:int,filename:str):
    ws = tools.open_xlsx_file(filename)
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title

    # 从老表中读出所有单元格数值,写入交换行列后索引后的新表
    for x in range(1,row+1):
        for y in range(1,column+1):
            new_ws.cell(row=y,column=x,value=ws.cell(row=x,column=y).value)

    new_wb_name = filename.split('.')[0]+'_rotate.xlsx'
    new_wb.save(new_wb_name)
    print(f'{new_wb_name}保存完成')

if __name__ == '__main__':
    rotate(5,4,'cell.xlsx')

