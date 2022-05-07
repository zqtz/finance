from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def fill_column(col_number: int, worksheet: Worksheet):
    """
    处理金蝶导出的凭证, 自动填充指定的列缺失的内容
    :param col_number: 要填充的列号
    :param worksheet: 要填充的工作表对象
    :return: 填充完成后的同一个工作表对象
    """
    current = 2
    max_index = worksheet.max_row - 1
    while current <= max_index:
        if not worksheet.cell(row=current, column=col_number).value:
            worksheet.cell(current, col_number, worksheet.cell(current - 1, col_number).value)

        current += 1
    return worksheet


if __name__ == '__main__':
    wb = load_workbook('kisdocument.xlsx')
    ws = wb.active
    fill_column(5, fill_column(6, ws))
    wb.save('new2.xlsx')
