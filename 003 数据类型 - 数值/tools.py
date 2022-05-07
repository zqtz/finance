import re
from openpyxl import load_workbook
import decimal
import datetime

def open_xlsx_file(file_name:str,sheet_name=None):
    if sheet_name:
        return load_workbook(file_name).active[sheet_name]
    return load_workbook(file_name).active


def transfer_to_decimal(num) -> decimal.Decimal:
    if type(num) == int:
        return decimal.Decimal(num)
    elif type(num) == float:
        return decimal.Decimal('{0:.2f}'.format(num))
    elif type(num) == str:
        if re.match('[+-]?\\d+(\\.\\d+)?$', num):
            if num.find('.') == -1:
                return decimal.Decimal(num)
            else:
                split_num = num.split('.')
                if len(split_num[1]) > 2:
                    num_string = split_num[0] + '.' + split_num[1][0:2]
                    if int(split_num[1][2]) >= 5:
                        if split_num[0][0] == '-':
                            return decimal.Decimal(num_string) - decimal.Decimal('0.01')
                        else:
                            return decimal.Decimal(num_string) + decimal.Decimal('0.01')
                    else:
                        return decimal.Decimal(num_string)
                elif len(split_num[1]) == 2:
                    num_string = split_num[0] + '.' + split_num[1][0:2]
                    return decimal.Decimal(num_string)
                elif len(split_num[1]) == 1:
                    num_string = split_num[0] + '.' + split_num[1] + '0'
                    return decimal.Decimal(num_string)
                else:
                    raise AttributeError

        else:
            return AttributeError

def read_excel(row,column):
    wb = load_workbook('datatypes.xlsx')
    ws = wb.active
    result = ws[f'{row}{column}'].value
    return result


def save_to_excel(column,row,data):
    wb = load_workbook('test.xlsx')
    ws = wb.active
    # ws['a1'] = 1
    # wb.save('test.xlsx')
    print(f'{column}{row}')
    ws[f'{column}{row}'] = data
    wb.save('test.xlsx')
    print('储存完成')

def get_datatype():
    wb = load_workbook('datatypes.xlsx')
    ws = wb.active
    for i in range(1,8):
        result = type(ws[f'C{i}'].value)
        print(result)

def get_sumifs():
    wb = load_workbook('amount.xlsx')
    ws = wb.active
    target_datetime = datetime.datetime(2017, 2, 7)
    sum_day = decimal.Decimal('0')
    for i in range(3,61):
        # 获取每个单元格的日期字符串
        currrent_date_string = ws['A'+str(i)].value
        #将这个日期字符串转化成datetime对象
        datetime_obj = datetime.datetime.strptime(currrent_date_string, '%Y%m%d')
        if datetime_obj.day == target_datetime.day and datetime_obj.month == target_datetime.month and datetime_obj.year == target_datetime.year:
            sum_day += transfer_to_decimal(ws[f'C{str(i)}'].value)
    print(sum_day)