import openpyxl
import decimal
from openpyxl import load_workbook
import re
import datetime
import tools

# wb = load_workbook('datatypes.xlsx')
# print(f'"{wb.active["A1"]}的数据类型为{type(wb.active["A1"].value)}')
# print(f'"{wb.active["A2"]}的数据类型为{type(wb.active["A2"].value)}')
# print(f'"{wb.active["A3"]}的数据类型为{type(wb.active["A3"].value)}')
# print(f'"{wb.active["A4"]}的数据类型为{type(wb.active["A4"].value)}')
def main():
    # print(transfer_to_decimal('100.'))
    # print(transfer_to_decimal('100.1'))
    # print(transfer_to_decimal('100.11'))
    # print(transfer_to_decimal('100.113'))
    # print(transfer_to_decimal('-100.115'))
    num_decimal = decimal.Decimal('0')
    wb = load_workbook('amount.xlsx')
    for i in range(3, 61):
        num_decimal += tools.transfer_to_decimal(wb.active['C' + str(i)].value)
    print(num_decimal)
    num_float = 0
    for i in range(3, 61):
        num_float += float(wb.active['C' + str(i)].value)
    print(num_float)


if __name__ == '__main__':
    # main()
    # get_datatype()
    # get_sumifs()
    data = [9223372036854775807,123.45,'12345',"字符串",datetime.datetime(1733,5,27,19),'=A1+A2']
    for j in range(1,len(data)+1):
        tools.save_to_excel('a',j,data[j-1])
    # save_to_excel()


