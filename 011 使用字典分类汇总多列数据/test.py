import openpyxl
import tools

def subtotal_single(key_column: int, value_column: int, file_name: str) -> dict:
    wb = tools.open_xlsx_file(file_name)

    result = {}

    # 2 -> wb.max_row -1 是要处理的列
    for i in range(2, wb.max_row):

        # 键是否存在于字典中
        key = wb.cell(row=i, column=key_column).value

        if key in result:

            # 存在的话, 需要更新
            result[key] = result[key] + tools.transfer_to_decimal(wb.cell(row=i, column=value_column).value)

        # 不存在的话, 直接设置
        else:
            result[key] = tools.transfer_to_decimal(wb.cell(row=i, column=value_column).value)

    return result

# 以一列为键, 合并汇总另外多列的值, 每个值是平行汇总的
def subtotal_composite(key_column: int, value_column1: int, value_column2: int, file_name: str) -> dict:
    wb = tools.open_xlsx_file(file_name)

    result = {}
    # 2 -> wb.max_row -1 是要处理的列
    for i in range(2, wb.max_row):

        # 键是否存在于字典中
        key = wb.cell(row=i, column=key_column).value

        if key in result:
            # 如果存在, 要更新两个值
            result[key][wb.cell(row=1, column=value_column1).value] = result[key][wb.cell(row=1,
                                                                                          column=value_column1).value] + tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column1).value)
            result[key][wb.cell(row=1, column=value_column2).value] = result[key][wb.cell(row=1,
                                                                                          column=value_column2).value] + tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column2).value)

        # 不存在的话, 创建键和对应的嵌套字典, 初始值是0
        else:
            result[key] = {}
            result[key][wb.cell(row=1, column=value_column1).value] = tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column1).value)
            result[key][wb.cell(row=1, column=value_column2).value] = tools.transfer_to_decimal(
                wb.cell(row=i, column=value_column2).value)

    return result

if __name__ == '__main__':
    # result = subtotal_single(4,7,'data.xlsx')
    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.cell(1, 1, '科目')
    # ws.cell(1, 2, '借方')
    # start = 2
    # for k,v in result.items():
    #     ws.cell(row=start,column=1,value=k)
    #     ws.cell(row=start, column=2, value=v)
    #     start +=1
    # wb.save('result.xlsx')

    result = subtotal_composite(4, 7, 8, 'data.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, '科目')
    ws.cell(1, 2, '借方')
    ws.cell(1, 3, '贷方')
    start = 2
    for k, v in result.items():
        ws.cell(row=start, column=1, value=k)
        ws.cell(row=start, column=2, value=v['借方'])
        ws.cell(row=start, column=3, value=v['贷方'])
        start += 1
    wb.save('result1.xlsx')





